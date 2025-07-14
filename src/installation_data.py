# This file defines installation-specific data for a customer/importer.
# It includes the "type of determination" indicator used in "prepare_data.py":
# - If type_of_determination = "02", default values are used for all entries (no installation-level data is collected).
# - If type_of_determination = "01", installation-level data is collected based on the status in the consultation_overview file.
#   Depending on this status, either default values or real data from supplierTool.xlsx are used.
#
# list of relevant installations is obtained from prepared data dict - only these are considered

import openpyxl
import pprint
import os
import shutil

# pdf creation
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from openpyxl.utils import column_index_from_string

from src.log import Log
from src.xlsx_access import load_table_to_dict, keyword_index, column_letter_from_string
from src.load_supplier_tool import load_supplier_tool, get_installation_folder_path
from src.helper import clean_supplier_str
from src.shared import shared_data


def get_installation_data(config, prepared_data, customer_dict):
    # type of determination * Test

    # extract a set of relevant installation/ operators
    relevant_installation_operators = set()
    for entry in customer_dict["Angaben_zu_Warenmengen"]["table_imported_goods"]:
        name = entry["operator_or_installation"]
        relevant_installation_operators.add(name)

    # derive determination type
    installation_data = {}

    determination_approach = get_det_approach(config, prepared_data)

    # < supporting_docs_adjustment 17/04 >
    if "default_supporting_docs" in config:
        installation_data["default_supporting_documents"] = config["default_supporting_documents"]
    # </>

    if determination_approach == "default_values":
        # -> All installations are using default_values
        # det_type "02"
        # Default values are determined within in prepare_data.py
        installation_data["fix_type_of_determination"] = "02"

        # if it's necessary to submit supporting documents with default values it must be added here

        shared_data_summary = shared_data["current"]["summary"]
        shared_data_summary["intern_report_type"] = "Default Values in Default Value Period"

        return installation_data

    elif determination_approach == "zero_report_with_default_docs":
        # creating zero report
        inst_entry = {}
        zero_report(inst_entry, prepared_data, config, None, zero_report_without_docs=False, default_supporting_docs=True)
        installation_data = {"default": inst_entry}
        installation_data["fix_type_of_determination"] = "03"  # is set in real_data_approach


        shared_data_summary = shared_data["current"]["summary"]
        shared_data_summary["intern_report_type"] = "Central: Zero Report with Default Documentation"

        return installation_data
    

    elif determination_approach == "zero_report_without_docs":

        inst_entry = {}
        zero_report(inst_entry, prepared_data, config, None, zero_report_without_docs=True, default_supporting_docs=True)

        installation_data = {"default": inst_entry}
        installation_data["fix_type_of_determination"] = "03"  # is set in real_data_approach

        shared_data_summary = shared_data["current"]["summary"]
        shared_data_summary["intern_report_type"] = "Central: Zero Report without Documentation"

        return installation_data

    elif determination_approach == "real_data_determination":
        # try to use real_data, otherwise zero report
        # accounts for det_type "01/03" in cases where determination is not possible

        # reads from ConsultationOverview.xlsx
        installation_overview = get_overview_data(config, prepared_data)

        if installation_overview is None:
            user_input = Log.dialog(
                f"No sheet corresponding to importer '<yellow>{prepared_data['general_info']['importer_name']}</r>' found in consultation overview file.\nDo you want to create a <b>zero report without documentation?</r>",
                title="consultation overview: customer not found",
            )

            if user_input == "y" or user_input == "yes":
                shared_data_summary = shared_data["current"]["summary"]
                shared_data_summary["intern_report_type"] = "Zero Report without Documentation"

                # creating report without supporting documents
                inst_entry = {}
                zero_report(inst_entry, prepared_data, config, None, zero_report_without_docs=True)
                installation_data = {"default": inst_entry}
                installation_data["fix_type_of_determination"] = "03"  # is set in real_data_approach
            else:
                os._exit(1)

        else:
            installation_data = real_data_approach(
                installation_overview,
                relevant_installation_operators,
                config,
                prepared_data,
            )
            installation_data["fix_type_of_determination"] = None  # is set in real_data_approach

        return installation_data


def real_data_approach(installation_overview, relevant_installation_operators, config, prepared_data):
    # constists of all installation entries
    installation_data = {}

    for installation_operator in relevant_installation_operators:
        installation_entry = None

        ## * Matching the installation data from the customer data with the overview file data

        for i in range(2):
            # search two times, second time in case installation_operator is an installation of form: "installation (operator)"

            # exact search
            for co_installation_entry in installation_overview:
                if co_installation_entry["installation"].strip().lower() == installation_operator.strip().lower():
                    if installation_entry is not None:
                        Log.warning(
                            f"\nFound installation {installation_operator} multiple times in overview file. Using the first one.",
                        )
                        continue
                    Log.debug(
                        f"Found exact match for installation {installation_operator} in overview file data",
                        key="installation_data",
                    )
                    installation_entry = co_installation_entry

                # installation not found in overview file

            # non-exact search
            if installation_entry is None:
                for co_installation_entry in installation_overview:
                    if clean_supplier_str(co_installation_entry["installation"].lower()) == clean_supplier_str(installation_operator.lower()):
                        if installation_entry is not None:
                            Log.warning(
                                f"Found multiple (non-exact) matches for installation {installation_operator}  in overview file. Using the first one.",
                                title="supplier_workflow | multiple installations",
                            )
                        Log.warning(
                            f"Found non-exact match for installation {installation_operator} in overview file data. Matched:\n'{installation_operator}'\n'{co_installation_entry['installation']}'",
                            title="supplier_workflow | non-exact match on installation",
                        )
                        installation_entry = co_installation_entry

            if installation_entry is not None:
                break

            if installation_entry is None and i == 0 and installation_operator.endswith(")"):
                # assume that entry is an installation, not an operator
                installation_operator = installation_operator.split("(")[0].strip()
                continue

            # throwing error if not found
            if installation_entry is None:
                Log.error(
                    f"Installation/ operator {installation_operator} not found in overview file data",
                    title="supplier_workflow | installation not found",
                )

            if i == 1:
                Log.info(f"Identified {installation_operator} as installation, not operator - found in overview file data")

        ## ** Determining the action to take, based on communication status

        try:
            action = config["supplier_workflow"]["status"][installation_entry["communication status"]]
        except Exception:
            Log.error(
                f"Installation/ operator {installation_operator} has no communication status set",
                title="supplier_workflow | No communication status",
            )

        if action == "abort":
            Log.error(
                f"Installation/ operator {installation_operator} has status '<orange>{installation_entry["communication status"]}</r>' in overview file data\n -> action: abort",
                title="supplier_workflow | action: abort",
            )

        elif action.startswith("zero_report"):
            # corresponds to "zero_report_create_docs", "zero_report_without_docs" and "zero_report_sup_docs"
            try:
                zero_report(installation_entry, prepared_data, config, action)
            except Exception as e:
                Log.error(
                    f"Error creating zero report for installation '{installation_entry['installation']}':\n{e}",
                    title="supplier_workflow | zero report error",
                )

            installation_entry["type_of_determination"] = "03"

            shared_data_summary = shared_data["current"]["summary"]
            shared_data_summary["intern_report_type"] = "Zero Report with Customer Documentation"

        elif action == "use_supplier_tool_data":
            # todo: this is faulty because it is visited once per installation or so
            shared_data_summary = shared_data["current"]["summary"]
            shared_data_summary["intern_report_type"] = "Real Data Report with Supplier Tool Data"

            installation_entry["type_of_determination"] = "01"

            try:
                load_supplier_tool(installation_entry, prepared_data, config)

            except Exception as e:
                Log.error(
                    f"Error loading supplier tool data for installation '{installation_entry['installation']}':\n{e}",
                    title="supplier_workflow | supplier tool data error",
                    delay_exit=True,
                )
                raise e
            

            # < supporting_docs_adjustment 17/04 >
            custom_additional_information, supporting_documents_list = get_supportings_documents_list(installation_entry, prepared_data, config)
            if supporting_documents_list is not None:
                installation_entry["supporting_documents"] = supporting_documents_list
                use_of_default_docs = config.get("use_default_supporting_docs", None)
                Log.info(
                    f"[Real Data Approach] Found {'<yellow> default </r>' if use_of_default_docs else ''} supporting documents for installation '{installation_entry['installation']}':\n{supporting_documents_list}, besides the supplier tool data given",
                )

        elif action == "ignore":
            Log.warning(
                f"Installation/ operator {installation_operator} has (unknown?) status '<orange>{installation_entry['communication status']}</r>' in overview file data\n -> action: ignore",
                title="supplier_workflow | action: ignore",
            )
            continue

        installation_data[installation_operator] = installation_entry

    return installation_data


def zero_report(installation_entry, prepared_data, config, action, zero_report_without_docs=False, default_supporting_docs=False):
    supporting_documents_list = None
    custom_additional_information = None

    # hier ansetzen falls supporting docs sowohl geladen als auch erzeugt werden müssen
    if default_supporting_docs:
        custom_additional_information, supporting_documents_list = get_supportings_documents_list(installation_entry, prepared_data, config)

    elif not zero_report_without_docs: 
        if action.endswith("sup_docs"):
            # * Load supporting documents from provided data
            custom_additional_information, supporting_documents_list = get_supportings_documents_list(installation_entry, prepared_data, config)

            if len(supporting_documents_list) > 0:
                Log.info(
                    f"<b>Found <green>{len(supporting_documents_list)}</r><b> supporting documents for installation '<green>{installation_entry['installation']}</r>': \n"
                    + "\n".join(f" - {'[...] ' if len(entry) > 60 else ''}{entry[-60:]}" for entry in supporting_documents_list)
                )
                
                
            else:
                Log.warning(
                    f"No supporting documents found for installation '<yellow>{installation_entry['installation']}</r>', despite action code '<yellow>{action}</r>'",
                    title="supplier_workflow | no supporting documents",
                )

        elif action.endswith("without_docs"):
            Log.warning(
                f"According to consultation_overview status 'X1':\nZero report without supporting documents for installation '<yellow>{installation_entry['installation']}</r>'",
                title="supplier_workflow | zero report",
            )
            zero_report_without_docs = True

        else:
            # * create supporting documents from consultation overview information

            Log.error("tried to create supporting documents from consultation overview information -> deprecated \n",
                title="supplier_workflow | zero report",
            )

            comm_attempt_entries = communication_attempt_entries(installation_entry)
            supporting_documents_list = [("<template>")]

    if custom_additional_information is None:
        additional_info = "Es war nicht möglich die Emissionsdaten zu ermitteln."

        if not zero_report_without_docs:
            additional_info += " Eine Dokumentation der gescheiterten Versuche zur Ermittlung der Daten"
            additional_info += " ist in folgenden Dateien unter 'Supplementary/Supporting documents' zu finden:"

            for doc in supporting_documents_list:
                additional_info += f"\n - {os.path.basename(doc)}"

    else:
        additional_info = custom_additional_information

        Log.info(
            f"<b>Found additional information for installation '<green>{installation_entry['installation']}</r>':\n<i>{additional_info}</r>",
        )

    if supporting_documents_list is not None:
        installation_entry["supporting_documents"] = supporting_documents_list

    # dict structure

    installation_entry["emission_data"] = {}
    installation_entry["emission_data"]["default"] = {}  # placeholder for every cn code
    emission_data_def = installation_entry["emission_data"]["default"]

    # direct emissions

    direct_emissions = emission_data_def["direct_emissions"] = {}
    direct_emissions["reporting_methodology"] = None  # Nicht angeben
    direct_emissions["additional_info"] = additional_info
    direct_emissions["see"] = 0
    direct_emissions["type_of_determination"] = "03"

    # direct emissions

    indirect_emissions = emission_data_def["indirect_emissions"] = {}
    indirect_emissions["type_of_determination"] = "03"

    indirect_emissions["source_of_emission_factor"] = None
    indirect_emissions["source_of_emission_factor_value"] = None
    indirect_emissions["other_source_indication"] = None
    indirect_emissions["source_of_electricity"] = None

    indirect_emissions["emission_factor"] = 0
    indirect_emissions["electricity_consumed"] = 0
    indirect_emissions["see"] = 0


# creates a pdf with the supporting document data
def create_supporting_documents(installation_entry, comm_attempt_entries, config):
    document_name = installation_entry["installation"].strip().replace(" ", "_").lower()
    document_name = f"{document_name}.pdf"

    # Speicherort
    output_dir = shared_data["current"]["output_dir"]
    output_dir = os.path.join(output_dir, "temp")
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    filepath = os.path.join(output_dir, document_name)

    # PDF-Dokument mit A4-Seitengröße erstellen
    doc = SimpleDocTemplate(filepath, pagesize=A4, rightMargin=2 * cm, leftMargin=2 * cm, topMargin=2 * cm, bottomMargin=2 * cm)

    # Elemente für das PDF
    elements = []

    # Stile für den Titel und den Text
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.alignment = 1  # Zentriert

    # Titel hinzufügen
    title = Paragraph("Dokumentation gescheiterter Kontaktversuche zur Ermittlung der Emissionsdaten", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.5 * cm))  # Abstand

    # Tabellenkopf und Daten für die Tabelle erstellen
    table_data = [["#", "Datum", "Anmerkungen"]]  # Tabellenkopf

    # Einträge aus comm_attempt_entries in die Tabelle hinzufügen
    for idx, entry in enumerate(comm_attempt_entries, start=1):
        # Anmerkungen kombinieren
        anmerkungen = f"{entry['remarks_dropdown']} - {entry['remarks']}"
        # Zeile hinzufügen
        table_data.append([str(idx), entry["date"], anmerkungen])

    # Tabelle erstellen und Spaltenbreiten festlegen
    table = Table(table_data, colWidths=[3 * cm, 4 * cm, 10 * cm])

    # Stil für die Tabelle
    style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgreen),  # Kopfzeile hellblassgrün
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),  # Schwarzer Text
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),  # Text linksbündig
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),  # Fettschrift für die Kopfzeile
            ("FONTSIZE", (0, 0), (-1, 0), 12),  # Größere Schrift in der Kopfzeile
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),  # Weißer Hintergrund für die Datenzeilen
            ("GRID", (0, 0), (-1, -1), 1, colors.black),  # Schwarzes Gitter um alle Zellen
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),  # Abstand unten in allen Zellen
            ("TOPPADDING", (0, 0), (-1, -1), 6),  # Abstand oben in allen Zellen
        ]
    )
    table.setStyle(style)

    # Tabelle hinzufügen
    elements.append(table)

    # Abstand hinzufügen
    elements.append(Spacer(1, 1 * cm))

    # Aktuelles Datum hinzufügen (unten rechts)
    current_date = datetime.now().strftime("%Y-%m-%d")
    footer = Paragraph(f"Erstellt am: {current_date}", styles["Normal"])
    elements.append(Spacer(1, 0.5 * cm))  # Etwas Platz vor dem Footer
    elements.append(footer)

    # PDF generieren
    doc.build(elements)

    # Rückgabe des Dateipfads
    return [filepath]


def communication_attempt_entries(installation_entry):
    empty_entries = 0
    no_textbox_entires = 0

    textbox_entries = []
    while f"date of attempt {no_textbox_entires + 1}" in installation_entry:
        # reading numbered "communication attempt" entries from the consultation overview file
        no_textbox_entires += 1
        date = installation_entry[f"date of attempt {no_textbox_entires}"]
        remarks = installation_entry[f"remarks {no_textbox_entires}"]
        remarks_dropdown = installation_entry[f"remarks dropdown {no_textbox_entires}"]
        remarks = "" if remarks is None else remarks
        remarks_dropdown = "" if remarks_dropdown is None else remarks_dropdown
        remarks = installation_entry[f"remarks {no_textbox_entires}"]

        if date == "" or date is None:
            empty_entries += 1
            continue

        textbox_entries.append(
            {
                "date": date,
                "remarks": remarks,
                "remarks_dropdown": remarks_dropdown,
            }
        )

    if no_textbox_entires - empty_entries < 2:
        Log.error(
            f"Installation/ operator {installation_entry['installation']} has only {no_textbox_entires - empty_entries} attempt(s) of communication and is a zero report without provided supporting documents.",
            title="supplier_workflow | zero report",
        )

    return textbox_entries


def get_supportings_documents_list(installation_entry, prepared_data, config):
    # todo: rename the files if necessary

    default_docs = config.get("default_supporting_documents", None)
    default_additional_information = config.get("default_additional_information", None)  # not implemented yet

    if default_docs is not None:
        
        # < supporting_docs_adjustment 17/04 >
        # copy file to temp folder
        temp_folder_path = os.path.join(shared_data["current"]["output_dir"], "temp")
        for index, filepath in enumerate(default_docs):
            file_name = os.path.basename(filepath)
            dst_path = os.path.join(temp_folder_path, file_name)
            os.makedirs(os.path.dirname(dst_path), exist_ok=True)
            shutil.copy(filepath, dst_path)
            # change file path to new filename
            #default_docs[index] = file_name
        # </>

        return default_additional_information, default_docs
        
    installation_folder_path = get_installation_folder_path(installation_entry, prepared_data, config)

    supporting_documents = []

    possible_folders = [
        f for f in os.listdir(installation_folder_path)
        if "supporting_document" in f.lower() and os.path.isdir(os.path.join(installation_folder_path, f))
    ]

    if possible_folders:
        sup_doc_folder_name = possible_folders[0]
        sup_doc_path = os.path.join(installation_folder_path, sup_doc_folder_name)
    else:
        sup_doc_path = None

    custom_additional_information_path = os.path.join(installation_folder_path, "custom_additional_information.txt")

    if os.path.exists(custom_additional_information_path):
        with open(custom_additional_information_path, "r") as f:
            custom_additional_information = f.read()
        # mark the file as read, by renaming it into "custom_additional_information - done.txt
        os.rename(custom_additional_information_path, custom_additional_information_path.replace(".txt", " - done.txt"))

    else:
        # if already renamed, check for the done file
        custom_additional_information_path = custom_additional_information_path.replace(".txt", " - done.txt")
        if os.path.exists(custom_additional_information_path):
            with open(custom_additional_information_path, "r") as f:
                custom_additional_information = f.read()
        else:
            custom_additional_information = None

    if sup_doc_path is None:
        Log.error(
            f"No supporting documents found in folder containing 'supporting_document' under '<yellow>{installation_folder_path}</r>'",
            title="Supporting Documents Missing"
        )

    if os.path.exists(sup_doc_path):
        # Liste der Dateien filtern, um versteckte Dateien auszuschließen und vollständige Pfade zu erhalten
        supporting_documents = [
            os.path.join(sup_doc_path, f)  # Vollständiger Pfad
            for f in os.listdir(sup_doc_path)
            if not f.startswith(".") and os.path.isfile(os.path.join(sup_doc_path, f))
        ]

        # copy file to temp folder

        temp_folder_path = os.path.join(shared_data["current"]["output_dir"], "temp")
        for filepath in supporting_documents:
            file_name = os.path.basename(filepath)
            dst_path = os.path.join(temp_folder_path, file_name)
            os.makedirs(os.path.dirname(dst_path), exist_ok=True)
            shutil.copy(filepath, dst_path)
            # change file path to new filename
            supporting_documents[supporting_documents.index(filepath)] = file_name
        
    return custom_additional_information, supporting_documents


def get_det_approach(config, prepared_data):
    general_info = prepared_data["general_info"]
    report_year = general_info["year"]
    report_quarter = general_info["quarter"]

    preset = config.get("preset_determination_approach", None)
    if preset is not None and preset != "" and preset != "None":
        Log.warning(
            f"Using pre-set determination approach: '<yellow>{config['preset_determination_approach']}</r>' for Q{report_quarter}-{report_year}",
            title="supplier_workflow | pre-set determination approach",
        )
        return config["preset_determination_approach"]

    preset = config.get("tagset_determination_approach", None)
    if preset is not None and preset != "" and preset != "None":
        Log.warning(
            f"Using tagset determination approach: '<yellow>{config['tagset_determination_approach']}</r>' for Q{report_quarter}-{report_year}",
            title="supplier_workflow | tagset determination approach",
        )
        return config["tagset_determination_approach"]

    action = config["determination_approach"][f"Q{report_quarter}-{report_year}"]

    return action


def get_overview_data(config, prepared_data):
    overview_file_path = config["supplier_workflow"]["consultation_overview_file"]
    general_info = prepared_data["general_info"]
    importer_name = general_info["importer_name"]

    # * load the consultation overview file

    try:
        wb = openpyxl.load_workbook(overview_file_path, data_only=True)
    except FileNotFoundError:
        Log.error(
            f"Error: Consultation overview file not found at {overview_file_path}",
            title="ConsultationOverview file not found",
        )
        return

    # * Load the alias names if given

    alias_head_field = keyword_index(wb["1 - Customer List"], "Alias", fixed_row=3)

    if alias_head_field is None:
        Log.error(
            "Error: Alias field not found in consultation overview file",
            title="supplier_workflow | alias field not found",
        )
        return

    real_name_field = keyword_index(wb["1 - Customer List"], "Real Name", fixed_row=3)

    if real_name_field is None:
        Log.error(
            "Error: Real Name field not found in consultation overview file",
            title="supplier_workflow | real name field not found",
        )
        return

    alias_dict = {}

    alias_col = column_index_from_string(column_letter_from_string(alias_head_field))
    real_name_col = column_index_from_string(column_letter_from_string(real_name_field))

    for row in wb["1 - Customer List"].iter_rows(min_row=4, max_row=wb["1 - Customer List"].max_row, min_col=alias_col, max_col=real_name_col):
        alias = row[0].value
        real_name = row[1].value

        if alias is None or real_name is None:
            continue

        if clean_supplier_str(real_name.lower()) == clean_supplier_str(importer_name.lower()):
            if alias is not None and alias.strip() != "":
                importer_name = alias

                Log.info(f"Found alias '{alias}' for importer '{real_name}' in consultation overview file")

    # * find the importer sheet within the excel file

    importer_sheet = None

    for sheet in wb.sheetnames:
        if sheet.lower() == importer_name.lower():
            importer_sheet = wb[sheet]
            break

    if importer_sheet is None:
        # No sheet found -> try to determine sheet with non exact approach
        for sheet in wb.sheetnames:
            if clean_supplier_str(sheet.lower()) == clean_supplier_str(importer_name.lower()):
                importer_sheet = wb[sheet]
                Log.warning(
                    f"Non-exact match on importer {importer_name} in sheet {sheet}, Matched: \n\n'{importer_name}'\n\n'{sheet}'",
                    title="supplier_workflow | non exact match on importer sheet",
                )
                break

    if importer_sheet is None:
        # Still no sheet found -> error
        Log.warning(
            f"Error: No importer sheet found for importer {importer_name}",
            title="supplier_workflow | importer sheet not found",
        )
        return None

    # * load table

    # * determine column indices

    columns = [
        "installation",
        "communication status",
        "date of last update",
        "operator",
        "date of attempt $x",
        "remarks dropdown $x",
        "remarks $x",
    ]

    number_of_comm_attempts = 8

    for col in columns[:]:  # iterate over copy of list
        if col.endswith("$x"):
            base = col[:-2]
            for idx_ in range(1, number_of_comm_attempts + 1):
                new_col = f"{base}{idx_}"
                columns.append(new_col)
            columns.remove(col)

    head_index_keyword = "installation"

    # * load the table into a dict

    return load_table_to_dict(importer_sheet, head_index_keyword, columns)


class Supplier_installation:
    inst_default_type_of_determination: str

    cn_code_data = {}
    # possible data points: type of determination, source of emission factor, source of electricity,
    # other source indication, electricity consumed, emission factor, source of emissions factor value,
    # production method, reducing agent, reducing agent for precursor, Mn content, Cr content, Ni content,
    # other alloy content, C content, steel scrap usage, steel pre-consumer scrap, steel product scrap usage,
    # non-iron content, aluminium scrap usage, pre-consumer scrap, non-aluminium content

    def __init__(self, inst_default_type_of_determination: str):
        self.inst_default_type_of_determination = inst_default_type_of_determination
