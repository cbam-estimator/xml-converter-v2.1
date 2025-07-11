# load_customer.py
"""

- load the input file from a string
- load it with openpyxl
- determine the version, either directly or autodetermine
- load layout_instructions from a .yml file

"""

import yaml
from yaml.loader import SafeLoader
import pprint
from .log import Log
from . import xlsx_access as xls
from . import validator as val
from . import helper


import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
import openpyxl  # noqa
from openpyxl.utils import column_index_from_string, get_column_letter  # noqa


def process_input_file(input_file, version_layouts_file):
    file_layout_dict = load_version_layouts(version_layouts_file)

    version = determine_version(input_file, file_layout_dict)
    Log.info(f"Detected version: {version}")

    version_layout_dict = inheret_version(file_layout_dict, version)

    return load_customer_sheet(input_file, version_layout_dict)


def load_customer_sheet(input_file, file_layout):
    """
    TODO : Full Support for legacy structure, not only with sheets
    """
    raw_data_dict = {}

    wb = openpyxl.load_workbook(input_file, data_only=True)

    for sheet in file_layout["sheets"]:
        # iterate through sheets

        if "code_sheet_name" in sheet:
            sheet_name_ls = sheet["alias_sheet_name"]
            code_sheet_name = sheet["code_sheet_name"]
        else:
            Log.error(f"[load_customer_sheet] Missing 'code_sheet_name' in sheet '{sheet}'")

        if type(sheet_name_ls) is not list:
            sheet_name_ls = [sheet_name_ls]

        ws = None
        for sheet_name in sheet_name_ls:
            if sheet_name in wb.sheetnames:
                # sheet name can also be a list
                ws = wb[sheet_name]
                break
        if ws is None:
            Log.warning(f"! Sheet '{sheet_name}' not found in input file")
            continue

        raw_data_sheet = {}
        raw_data_dict[code_sheet_name] = raw_data_sheet

        Log.debug(f"Processing sheet '{sheet_name}'", key="load_customer_sheet")

        for table in sheet["tables"]:
            # iterate through tables in worksheets

            table_name = table["table_name"]
            Log.debug(f"- Processing table '{table_name}'", key="load_customer_sheet")

            raw_data_table = []
            raw_data_sheet[table_name] = raw_data_table

            num_entries = table["num_entries"]
            num_entries = int(num_entries.replace(".", ""))

            upper_left = table["upper_left"]

            if not isinstance(upper_left, list):
                upper_left = [upper_left]

            # upper_left is a list now

            upper_left_index = None

            for i, ul in enumerate(upper_left):
                if ul.startswith("auto:"):
                    upper_left_index = xls.keyword_index(ws, ul[5:])
                else:
                    upper_left_index = xls.keyword_index(ws, ul)

                if upper_left_index is None:
                    Log.debug(
                        f" (list) Could not determine upper left cell for table '{table_name}' in sheet '{sheet_name}' / (upper_left = {upper_left}, table['upper_left'] = {table['upper_left']})",
                        key="load_customer_sheet",
                    )
                    continue

                else:
                    break

            if upper_left_index is None:
                Log.error(
                    f"Could not determine upper left cell for table '<blue>{table_name}</blue>' in sheet '<blue>{sheet_name}</blue>' / (upper_left = <blue>{upper_left}</blue>, table['upper_left'] = <blue>{table['upper_left']}</blue>)"
                )

            upper_left_index, is_valid = val.process_and_validate("cell_index", upper_left_index)

            if not is_valid:
                Log.error(
                    f"Invalid upper left index '<blue>{upper_left_index}</blue>' in sheet '{sheet_name}' : table '{table_name}'",
                    title="Invalid upper left index",
                )

            primary_key = None
            primary_key_candidate = None
            head_row = "".join(filter(str.isdigit, upper_left_index))
            head_col = "".join(filter(str.isalpha, upper_left_index))

            ignore_fields = []
            for field in table["fields"]:
                # iterate through fields in tables to collect indices

                if "code_field_name" not in field:
                    Log.error(f"[load_customer_sheet] Missing 'code_field_name' in field '{field}' in table '{table_name}' in sheet '{sheet_name}'")

                code_field_name = field["code_field_name"]

                if "alias_field_name" in field:
                    alias_field_names = field["alias_field_name"]
                else:
                    alias_field_names = [code_field_name]

                if table["orientation"] == "vertical":
                    # for vertical tables, head index is a column letter
                    # note that alias_field_name can also be a list

                    required = True if "required" not in field or field["required"] else False

                    head_index_field = xls.keyword_index(ws, alias_field_names, fixed_row=head_row, required=required)

                    if head_index_field is None and not required:
                        ignore_fields.append(code_field_name)  # add to delete from table["fields"]
                        continue

                    field["head_index"] = "".join(filter(str.isalpha, head_index_field))

                elif table["orientation"] == "horizontal":
                    # for horizontal tables, head index is a row number
                    head_index_field = xls.keyword_index(ws, alias_field_names, fixed_column=head_col, required=True)
                    field["head_index"] = "".join(filter(str.isdigit, head_index_field))

                if "primary_key" in field and field["primary_key"]:
                    # the primary key is the row or column that is used to identify the entry
                    primary_key = str(field["head_index"])

                if primary_key_candidate is None and field["type"].endswith("_m"):
                    # primary_key_candidate is the first mandatory field and is used if no primary key is specified
                    primary_key_candidate = str(field["head_index"])

            for field in ignore_fields:
                table["fields"] = [f for f in table["fields"] if f["code_field_name"] != field]

            if primary_key is None:
                if primary_key_candidate is not None:
                    primary_key = primary_key_candidate
                else:
                    Log.error(f"[load_customer_sheet] Could not determine primary key for '{table_name}' in sheet '{sheet_name}'")

            report_empty = False
            num_actual_entries = 0

            offset = 1
            if "num_examples" in table and table["num_examples"] is not None:
                offset = int(table["num_examples"]) + 1

            for i in range(offset, num_entries + offset):
                # iterate through entries (rows or columns) in tables

                if table["orientation"] == "vertical":
                    pk_field = primary_key + str(i + int(head_row))
                    primary_key_val = ws[pk_field].value
                elif table["orientation"] == "horizontal":
                    pk_field = get_column_letter(i + column_index_from_string(head_col)) + primary_key
                    primary_key_val = ws[pk_field].value
                if primary_key_val is None or primary_key_val == "" or primary_key_val == "--":
                    report_empty = True
                    continue
                else:
                    num_actual_entries += 1

                if report_empty:
                    Log.warning(
                        f"[load_customer_sheet] Empty row encountered in table '{table_name}' before row {i + int(head_row)} \n primary_key = {primary_key} ; primary_key_val = {primary_key_val}"
                    )
                    report_empty = False

                entry = {}

                for field in table["fields"]:
                    # iterate for fields for given row or column

                    if "code_field_name" in field and "alias_field_name" in field:
                        alias_field_name = field["alias_field_name"]
                        code_field_name = field["code_field_name"]
                    else:
                        alias_field_name = code_field_name = field["field_name"]

                    if table["orientation"] == "vertical":
                        index = field["head_index"] + str(i + int(head_row))
                        value = ws[index].value
                    elif table["orientation"] == "horizontal":
                        index = get_column_letter(i + column_index_from_string(head_col)) + field["head_index"]
                        value = ws[index].value

                    value = "" if value is None else value
                    valid = True

                    try:
                        value, valid = val.process_and_validate(
                            field["type"],
                            value,
                            entry_data=entry,
                            condition=field.get("condition"),
                        )

                    except Exception as e:
                        Log.error(
                            f"[Con.][load_customer_sheet] Error cleaning and validating value for type {field['type']} in sheet '{sheet_name}' : '{table_name}' : '{code_field_name + "/" + alias_field_name}' [{index}]\n error: {e}"
                        )
                        raise e

                    if not valid:
                        Log.error(
                            f"""invalid value '<blue>{value}</r>' for val-type '<purple>{field["type"]}</purple>'.

Path: {{ '<blue>{sheet_name}</blue>' : '<blue>{table_name}</blue>' : '<blue>{code_field_name + "/" + alias_field_name}</blue>' [<blue>{field['head_index'] + str(i + int(head_row))}</blue>]

Head index {field['head_index']} ; i = {i} ; head_row = {head_row} }}
                            """,
                            title="Invalid Value",
                        )

                    entry[code_field_name] = value

                # edit  entry (row or column)



                entry["pk_index"] = pk_field

                raw_data_table.append(entry)

            if num_actual_entries == 0:
                Log.warning(f"[load_customer_sheet] Empty table '{table_name}' in sheet '{sheet_name}'")

    return raw_data_dict


def determine_version(input_file, version_layout_file):
    wb = openpyxl.load_workbook(input_file)
    spec_sheet_name = "Specification"

    if spec_sheet_name in wb.sheetnames:
        ws = wb[spec_sheet_name]
        version = xls.cell_adjacent_to_keyword(
            ws,
            ["Version", "Version:"],
            "right",
            case_sensitive=True,  # This might be 'Version:' somewhere
        )

        _, valid = val.process_and_validate("version", version)
        if not valid:
            raise ValueError(f"Invalid version '{version}' in sheet '{spec_sheet_name}'")

        version_str = f"version_{version.replace('.', '_')}"

        if version_str in version_layout_file:
            return version_str
        else:
            Log.error(f"Version '{version}' / '{version_str}' not found in version layout file")

    else:
        # try to auto-detect version

        # VERSION 1.7 - with installations but without inward processing infos

        ws = wb["Angaben_zu_Warenmengen"]

        inward_processing_does_not_exist = (
            xls.keyword_index(ws, "Ursprünglich zur Veredelung importiert", required=False) is None
            and xls.keyword_index(ws, "Menge noch nicht veredelter Ware...", required=False) is None
            and xls.keyword_index(ws, "Zulassungsstaat", required=False) is None
        )

        wb_installations_exists = "Produktions_Standorte_Liste" in wb.sheetnames

        if inward_processing_does_not_exist and wb_installations_exists:
            return "version_1_7"
        else:
            Log.debug(
                f"Not version 1.7:\ninward_processing_does_not_exist = {inward_processing_does_not_exist}\nwb_installations_exists = {wb_installations_exists}\n worksheets:{wb.sheetnames}",
                key="determine_version",
            )

        Log.warning("Could not determine version from input file. Using 1.7.2 (base) as default.")
        return "version_1_7_2"


def load_version_layouts(version_layout_file):
    layout_dict = None

    try:
        with open(version_layout_file, "r", encoding="utf-8") as f:
            layout_dict = yaml.load(f, Loader=SafeLoader)
    except FileNotFoundError:
        Log.error("[load_version_layouts] The specified .yml file was not found.")
    except yaml.YAMLError as exc:
        Log.error(f"[load_version_layouts] Error parsing YAML file: {exc}")
    except Exception as e:
        Log.error(f"[load_version_layouts] An unexpected error occurred: {e}")
    if not layout_dict:
        Log.error("[load_version_layouts] No data loaded from the YAML file.")

    layout_dict = helper.convert_dict_to_lowercase(layout_dict, protect_keys=["code_sheet_name", "alias_sheet_name"])

    # Log.procedure(f"Loaded version layouts from {version_layout_file}")

    return layout_dict


def inheret_version(layout_dict, version):
    """
    Versionen können Arbeitsblattnamen überschreiben und gesamte sheets ändern
    legacy name muss dabei gleich bleiben
    """

    base_dict = layout_dict["base"]
    version_dict = layout_dict[version]

    if "sheets" not in version_dict:
        Log.warning(f"[inheret_version] No 'sheets' key found in specified version '{version}' layout")
        return base_dict

    for sheet in version_dict["sheets"]:
        version_code_name = sheet["code_sheet_name"]

        for i, base_sheet in enumerate(base_dict["sheets"]):
            if "code_sheet_name" not in base_sheet:
                Log.error(f"[inheret_version] Sheet '{base_sheet['sheet_name']}' in base layout does not have a code_sheet_name")

            # if the sheet is present in the version layout dictionary, overwrite the whole base sheet object
            if base_sheet["code_sheet_name"] == version_code_name:
                base_dict["sheets"][i] = sheet  # Hier wird das Original-Objekt in base_dict geändert
                Log.debug(
                    f"sheet {version_code_name} was changed within base layout: {base_sheet}\n\n\n {base_dict}",
                    key="inherit_version",
                )

    return base_dict
