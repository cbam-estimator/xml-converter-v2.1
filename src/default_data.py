import openpyxl
from . import xlsx_access as xlsx
from . import validator as v
from .log import Log
from openpyxl.utils import column_index_from_string, get_column_letter


class DefaultData:
    data_dict = None
    is_initialized = False

    @staticmethod
    def initialize(config):
        DefaultData._load(config)
        DefaultData.is_initialized = True

    @staticmethod
    def get(key):
        DefaultData.init_check()
        return DefaultData.data_dict[key]

    @staticmethod
    def init_check():
        if not DefaultData.is_initialized:
            Log.error("Default data class called before initialization")

    @staticmethod
    def _load(config):
        file = config["default_data_file"]
        wb = openpyxl.load_workbook(file)

        default_data = {}

        # * options from the config file

        options = config["options"]
        default_data["options"] = options

        # * Our custom default values for the report

        ws = wb["report_default_values"]

        default_data["report_default_values"] = {
            "Declarant.IdentificationNumber": xlsx.cell_adjacent_to_keyword(
                ws, "Declarant.IdentificationNumber"
            )
        }

        # * CN Code default values from the EC

        ws = wb["cn_code_default_values"]

        columns = {
            "cn_code": "cn_code_m",
            "see_direct": "float_m",
            "see_indirect": "float_m",
            "description_of_goods": "string_m",
            "PM1": "string_m",
            "PM2": "string_o",
            "PM3": "string_o",
        }

        cn_code_data = DefaultData._load_default_table(ws, columns, "CN Code")

        cn_code_data = {entry["cn_code"]: entry for entry in cn_code_data}

        default_data["cn_code_default_values"] = cn_code_data

        # * Country default values

        ws = wb["country_default_values"]

        columns = {
            "country_code": "country_code",
            "english_name": "string",
            "german_name": "string",
            "alias1": "string",
            "alias2": "string",
        }

        country_data = DefaultData._load_default_table(ws, columns, "country_code")

        country_db = CountryDB()

        for entry in country_data:
            country_db.add_country(
                entry["english_name"],
                entry["german_name"],
                [entry["alias1"], entry["alias2"]],
                entry["country_code"],
            )

        default_data["country_data"] = country_db

        DefaultData.data_dict = default_data

    @staticmethod
    def _load_default_table(ws, columns, upper_left_str):
        upper_left = xlsx.keyword_index(ws, upper_left_str)

        start_row = int("".join(filter(str.isdigit, upper_left))) + 1
        start_column = column_index_from_string(
            "".join(filter(str.isalpha, upper_left))
        )

        data = []

        table_end = False
        faulty_entry_num = 0

        for row in ws.iter_rows(min_row=start_row):
            entry = {}

            faulty_entry = False

            for cell in row:
                if cell.column < start_column:
                    continue

                i = cell.column - start_column

                # table nend
                if i == 0 and cell.value is None:
                    table_end = True
                    break

                # row end
                if i >= len(columns):
                    break

                key = list(columns.keys())[i]

                Log.mute()
                val, valid = v.process_and_validate(
                    columns[key], cell.value, raise_errors=False, mute=True
                )
                Log.unmute()

                if not valid:
                    Log.debug(
                        f"[load_default]\nInvalid value '{cell.value}' in column '{key}' / index: {get_column_letter(cell.column) + str(cell.row)} in sheet '{ws.title}'\n → Skipping row...",
                        "load_default",
                    )
                    faulty_entry = True
                    faulty_entry_num += 1
                    break

                entry[key] = val

            if table_end:
                break
            if not faulty_entry:
                data.append(entry)

        if faulty_entry_num > 0:
            Log.debug(
                f"[load_default]\nSkipped {faulty_entry_num} faulty entries in sheet '{ws.title}'",
                key="load_default",
            )

        return data


class CountryDB:
    def __init__(self):
        self.english_name_dict = {}
        self.german_name_dict = {}
        self.alias_dict = {}
        self.un_code_dict = {}

    def add_country(self, english_name, german_name, aliases, un_code):
        # Die Struktur des Landes als dict speichern

        for alias in aliases[:]:  # Erstelle eine Kopie der Liste
            if alias is None:
                aliases.remove(alias)

        country_info = {
            "english_name": english_name,
            "german_name": german_name,
            "aliases": aliases,
            "un_code": un_code,
        }

        # Länderinformationen in den entsprechenden Dictionaries speichern
        self.english_name_dict[english_name.lower()] = country_info
        self.german_name_dict[german_name.lower()] = country_info
        self.un_code_dict[un_code] = country_info
        for alias in aliases:
            if alias is None:
                continue
            self.alias_dict[alias.lower()] = country_info

    def find_country(self, query):
        # Suche nach dem Land in allen Dictionaries
        query = query.lower()
        return (
            self.english_name_dict.get(query)
            or self.german_name_dict.get(query)
            or self.alias_dict.get(query)
            or self.un_code_dict.get(query)
        )

    def __repr__(self):
        countries_list = []
        for info in self.english_name_dict.values():
            country_repr = (
                f"English: {info['english_name']}, "
                f"German: {info['german_name']}, "
                f"Aliases: {', '.join(str(info['aliases']))}, "
                f"UN Code: {info['un_code']}"
            )
            country_repr = "<cntry> "
            countries_list.append(country_repr)
        return ":".join(countries_list)
