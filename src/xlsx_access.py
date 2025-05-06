# xlsx_access.py

from openpyxl.utils import get_column_letter, column_index_from_string
from .log import Log


def cell_adjacent_to_keyword(
    worksheet, keywords, orientation="right", case_sensitive=False
):
    """
    Returns the value of the cell adjacent to the first matching keyword found in the worksheet.

    Parameters:
    - worksheet: The worksheet object to search within.
    - keywords: A keyword (string) or a list of keywords to search for.
    - orientation: The direction to look for the adjacent cell ('right', 'left', 'up', 'down'). Default is 'right'.
    - case_sensitive: Boolean indicating whether the keyword search is case-sensitive. Default is False.

    Returns:
    - The value of the cell adjacent to the first matching keyword.
    - Returns an empty string "" if the target cell is empty.
    - Returns None if the target cell does not exist.

    Note:
    - Logs an error if none of the keywords are found in the worksheet.
    - Keeps all log, debug, and error messages consistent with the original function.
    """

    if orientation not in ["right", "left", "up", "down"]:
        raise ValueError(
            "Invalid orientation. Must be one of 'right', 'left', 'up', 'down'"
        )

    # Ensure keywords is a list
    if isinstance(keywords, str):
        keywords = [keywords]

    if not case_sensitive:
        keywords = [kw.lower() for kw in keywords]

    return_val = None
    matched_keyword = None

    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                c_value = cell.value.lower() if not case_sensitive else cell.value
            else:
                c_value = cell.value
            if c_value in keywords:
                matched_keyword = cell.value if case_sensitive else c_value
                if orientation == "right":
                    return_val = cell.offset(row=0, column=1).value
                elif orientation == "left":
                    if cell.column == 1:
                        raise ValueError(
                            "Cannot get cell to the left of the first cell in a row"
                        )
                    return_val = cell.offset(row=0, column=-1).value
                elif orientation == "up":
                    if cell.row == 1:
                        raise ValueError("Cannot get cell above the first row")
                    return_val = cell.offset(row=-1, column=0).value
                elif orientation == "down":
                    return_val = cell.offset(row=1, column=0).value
                break  # Stop searching after finding the first match
        if return_val is not None:
            break

    if return_val is None:
        Log.error(
            f"Could not find keyword '<yellow>{keywords}</r>' in worksheet '<yellow>{worksheet.title}</r>'",
            title="Could not find keyword",
        )
        return ""
    else:
        Log.debug(
            f"Found cell adjacent to keyword '{matched_keyword}' in worksheet {worksheet.title}: '<yellow>{return_val}</r>'",
            key="xlsx_access",
        )
        return return_val


def keyword_index(
    worksheet,
    keywords,
    fixed_row=None,
    fixed_column=None,
    required=False,
    case_sensitive=False,
    non_exact_supplier_search=False,
):
    """
    Returns the index of row and column of the keyword in the worksheet as a string.

    If fixed_row or fixed_column are provided, the search is limited to that row or column.

    Multiple keywords possible, will return the first found keyword information.
    Returns None if the keyword is not found in the worksheet or throws error if no keyword found and
    query marked as keyword required.
    """

    if type(keywords) is str:
        keywords = [keywords]

    if non_exact_supplier_search and case_sensitive:
        Log.error(
            "Non-exact supplier search requires case_sensitive=False",
            title="Invalid keyword_index arguments",
        )

    for index, keyword in enumerate(keywords):
        if not case_sensitive:
            keyword = keyword.lower()
        if non_exact_supplier_search:
            keyword = clean_str_for_supplier(keyword)

        keyword_ends_with_dots = keyword.endswith("...")

        if fixed_row is not None:
            if type(fixed_row) is str:
                fixed_row = int(fixed_row)
            for cell in worksheet[fixed_row]:
                value = cell.value.strip() if type(cell.value) is str else cell.value
                if value is None:
                    continue

                value = (
                    value.lower()
                    if type(value) is str and not case_sensitive
                    else value
                )

                if non_exact_supplier_search:
                    value = clean_str_for_supplier(value)

                if (
                    keyword_ends_with_dots
                    and isinstance(value, str)
                    and value.startswith(keyword[:-3])
                ) or value == keyword:
                    return get_column_letter(cell.column) + str(cell.row)
        elif fixed_column is not None:
            for cell in worksheet[fixed_column]:
                value = cell.value.strip() if type(cell.value) is str else cell.value
                if value is None:
                    continue

                value = (
                    value.lower()
                    if type(value) is str and not case_sensitive
                    else value
                )

                if non_exact_supplier_search:
                    value = clean_str_for_supplier(value)

                if (
                    keyword_ends_with_dots
                    and isinstance(value, str)
                    and value.startswith(keyword[:-3])
                ) or value == keyword:
                    return get_column_letter(cell.column) + str(cell.row)

        else:
            for row in worksheet.iter_rows():
                for cell in row:
                    value = (
                        cell.value.strip() if type(cell.value) is str else cell.value
                    )
                    value = (
                        value.lower()
                        if type(value) is str and not case_sensitive
                        else value
                    )
                    if (
                        keyword_ends_with_dots
                        and isinstance(value, str)
                        and value.startswith(keyword[:-3])
                    ) or value == keyword:
                        return get_column_letter(cell.column) + str(cell.row)

        if required and index == len(keywords) - 1:
            Log.error(
                f"Keywordlist <blue>{keywords}</blue> not found in worksheet <blue>{worksheet.title}</blue> <b>(required=True)</b>",
                title="Keyword / Column Header not found",
            )

        Log.debug(
            f"[keyword_index] Keywordlist '{keywords}' not found in worksheet {worksheet.title} (required=False)",
            key="load_default",
        )
    return None


def load_table_to_dict(worksheet, head_column, columns, orientation="vertical"):
    """
    Given a head column and columns, loads a table into a list of dictionaries (rows).
    Supports both vertical and horizontal orientations.
    Skips empty entries, indicated by an empty head column cell.

    Parameters:
    - worksheet: The worksheet object to read from.
    - head_column: The header column (for vertical orientation) or header row (for horizontal orientation).
    - columns: List of columns (for vertical) or rows (for horizontal) to extract.
    - orientation: 'vertical' (default) or 'horizontal', specifies the orientation of the table.

    Returns:
    - List of dictionaries containing the table data.

    Note:
    - Logs errors and warnings if the head column or specified columns/rows are not found.
    - Keeps all log, debug, and error messages consistent with the original function.
    """

    if orientation not in ["vertical", "horizontal"]:
        raise ValueError("Invalid orientation. Must be 'vertical' or 'horizontal'")

    column_indices = {}

    head_index = keyword_index(worksheet, head_column)

    if head_index is None:
        Log.error(
            f"Could not find head column '<yellow>{head_column}</r>' in '<yellow>{worksheet.title}</r>'",
            title="load_table_to_dict | head column not found",
        )

    head_index_row = row_index_from_string(head_index)
    head_index_col = column_letter_from_string(head_index)

    data = []

    if orientation == "vertical":
        # Existing behavior for vertical tables
        for column_name in columns:
            column = keyword_index(worksheet, column_name, fixed_row=head_index_row)
            if column is not None:
                column_indices[column_name] = column_letter_from_string(column)
            else:
                Log.warning(
                    f"Could not find column '{column_name}' in worksheet '{worksheet.title}'",
                    title="load_table_to_dict | column not found",
                )

        for row in worksheet.iter_rows(min_row=head_index_row + 1):
            entry = {}
            for column_name, column_index in column_indices.items():
                col_index_int = column_index_from_string(column_index) - 1
                value = row[col_index_int].value
                if column_name == head_column and value is None:
                    entry = {}
                    break
                entry[column_name] = value
            if entry:
                data.append(entry)

    elif orientation == "horizontal":
        # Adjusted behavior for horizontal tables
        row_indices = {}
        for row_name in columns:
            cell_ref = keyword_index(worksheet, row_name, fixed_column=head_index_col)
            if cell_ref is not None:
                row_number = row_index_from_string(cell_ref)
                row_indices[row_name] = row_number
            else:
                Log.warning(
                    f"Could not find row '{row_name}' in {worksheet.title}",
                    title="load_table_to_dict | row not found",
                )

        for col in worksheet.iter_cols(
            min_col=column_index_from_string(head_index_col) + 1
        ):
            entry = {}
            head_cell = col[head_index_row - 1]
            if head_cell.value is None:
                continue  # Skip columns with empty head cell
            for row_name, row_number in row_indices.items():
                value = col[row_number - 1].value
                if row_name == head_column and value is None:
                    entry = {}
                    break
                entry[row_name] = value
            if entry:
                data.append(entry)

    return data


# -> move to installation data and work with local import (ask chat, he knows)
def clean_str_for_supplier(supplier_str):
    """
    Cleans installation or operator string for comparison within supplier tool
    """

    if supplier_str is None:
        return None

    ignore = (
        " ",
        "-",
        "_",
        "(",
        ")",
        ".",
        ",",
        ";",
        ":",
        "/",
        "\\",
        "co",
        "ltd",
        "gmbh",
    )

    for char in ignore:
        supplier_str = supplier_str.replace(char, "")

    return supplier_str


def row_index_from_string(index_str):
    return int("".join(filter(str.isdigit, index_str)))


def column_letter_from_string(index_str):
    return "".join(filter(str.isalpha, index_str))
