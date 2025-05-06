import pprint
from .log import Log
import openpyxl
from openpyxl.utils import get_column_letter
from .xlsx_access import keyword_index, row_index_from_string, column_letter_from_string


def get_supplier_data(cn_code, general_info, operator_data, installation_data, config):
    target_importer = general_info["importer_name"]
    target_eori = general_info["importer_eori"]

    target_operator = operator_data["operator_name"]

    if installation_data is not None:
        target_name = installation_data["installation_name"]
        target_country = installation_data["installation_country"]
        target_city = installation_data["installation_city"]
        target_postcode = installation_data["installation_postcode"]
    else:
        target_name = operator_data["operator_name"]
        target_country = operator_data["operator_country"]
        target_city = operator_data["operator_city"]
        target_postcode = operator_data["operator_postcode"]

        Log.debug(
            f"No installation data provided for operator {target_name}",
            title=f"{cn_code} | {target_name}",
            key="supplier_data",
        )

    # find consultation overview

    cons_ow_path = config["supplier_workflow"]["consultation_overview_file"]

    # find importer worksheet within the excel file

    try:
        wb = openpyxl.load_workbook(cons_ow_path, data_only=True)
    except FileNotFoundError:
        Log.error(
            f"Error: Consultation overview file not found at {cons_ow_path}",
            title="ConsultationOverview file not found",
        )
        return

    importer_sheet = None

    for sheet in wb.sheetnames:
        if sheet.lower() == target_importer.lower():
            importer_sheet = wb[sheet]
            break

    if importer_sheet is None:
        Log.error(
            f"Error: No importer sheet found for operator {target_name}",
            title=f"{cn_code} | {target_name}",
            delay_exit=True,
        )
        return None

    # load table

    head_index_keyword = "installation"

    head_index_str = keyword_index(importer_sheet, head_index_keyword)
    head_index_row = row_index_from_string(head_index_str)
    head_index_col = column_letter_from_string(head_index_str)
    fields = {
        "installation": None,
        "communication status": None,
        "date of last update": None,
        "operator": None,
    }

    for field_name in fields.keys():
        idx = keyword_index(importer_sheet, field_name, fixed_row=head_index_row)
        if idx is not None:
            fields[field_name] = idx
        else:
            Log.warning(
                f"Could not find field '{field_name}' in supplier data",
                title=f"{cn_code} | {target_name}",
            )

    # find the row with the correct installation name

    keywords = [target_name, target_operator]

    if target_name is None or target_operator is None:
        Log.error(
            f"Error: No target name or operator provided for {cn_code}",
            title=f"{cn_code} | {target_name}",
            delay_exit=False,
        )
        return None

    installation_index = keyword_index(
        importer_sheet, keywords, fixed_column=head_index_col
    )

    if installation_index is None:
        installation_index = keyword_index(
            importer_sheet,
            keywords,
            fixed_column=head_index_col,
            non_exact_supplier_search=True,
        )

        if installation_index is not None:
            cell_value = importer_sheet[installation_index].value
            Log.warning(
                f"\nWarning! Settled with non-exact match in supplier data. Matched:\n\n'<orange>{target_name}</r>'\n\n'<orange>{cell_value}</r>'\n",
                title=f"Supplier Data / Non-exact match | {cn_code} | {target_name}",
            )

    if installation_index is None:
        Log.error(
            f"Error: No row found for target {target_name} / operator {target_operator}",
            title=f"{cn_code} | {target_name}",
            delay_exit=True,
        )
        return None

    cell_value = importer_sheet[installation_index].value
    Log.debug(
        f"Supplier Data / Found row for {target_name} / {target_operator} at index {installation_index} with value '{cell_value}'",
        title=f"{cn_code} | {target_name}",
        key="supplier_data",
    )

    # load data

    installation_info = {}
    field_row = str(row_index_from_string(installation_index))

    for field_name, field_index in fields.items():
        if field_index is None:
            installation_info[field_name] = None
        else:
            field_col = column_letter_from_string(field_index)
            installation_info[field_name] = importer_sheet[field_col + field_row].value

    installation_name = installation_info["installation"]
    communication_status = installation_info["communication status"]

    try:
        action = config["supplier_workflow"]["status"][communication_status]
    except KeyError:
        Log.error(
            f"Invalid status '<orange>{communication_status}</r>'   -   <orange>{target_importer}</r>  /  <orange>{installation_name}</r>",
            title=f"SupplierData | Invalid Status | {target_importer}",
        )

    if action == "abort":
        Log.error(
            f"Status '<orange>{communication_status}</r>' → {action}   -   <orange>{target_importer}</r>  /  <orange>{installation_name}</r>",
            title=f"SupplierData | Status abort | {target_importer}",
            delay_exit=False,
        )

    if action == "use_default":
        return "use_default"

    return {"see_direct": 0.0, "see_indirect": 0.0, "description_of_goods": ""}
